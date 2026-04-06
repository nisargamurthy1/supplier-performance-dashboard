"""
build_dashboard.py
Supplier Performance Dashboard
- Seeds SQLite database with realistic sample data
- Runs all SQL queries
- Exports a fully formatted Excel dashboard with 6 sheets
Author: Nisarga Narasimhamurthy
"""

import sqlite3
import pandas as pd
import random
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint

random.seed(42)

# ── CONFIG ───────────────────────────────────────────────────────────────────
DB_PATH    = "supplier_data.db"
OUT_PATH   = "supplier_performance_dashboard.xlsx"
TODAY      = datetime(2026, 4, 4)

# ── COLORS ───────────────────────────────────────────────────────────────────
DARK_NAVY  = "1a1a2e"
MID_NAVY   = "16213e"
BLUE       = "0F3460"
ACCENT     = "378ADD"
LIGHT_BLUE = "E6F1FB"
GREEN_BG   = "E1F5EE"
GREEN_TXT  = "0F6E56"
AMBER_BG   = "FAEEDA"
AMBER_TXT  = "854F0B"
RED_BG     = "FCEBEB"
RED_TXT    = "A32D2D"
WHITE      = "FFFFFF"
GRAY_HDR   = "F1EFE8"
GRAY_BDR   = "D3D1C7"

def hdr_font(sz=11, bold=True, color=WHITE):
    return Font(name="Arial", size=sz, bold=bold, color=color)

def body_font(sz=10, bold=False, color="2C2C2A"):
    return Font(name="Arial", size=sz, bold=bold, color=color)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def border():
    s = Side(style="thin", color=GRAY_BDR)
    return Border(left=s, right=s, top=s, bottom=s)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

# ── SEED DATABASE ─────────────────────────────────────────────────────────────
def seed_database():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()

    c.executescript("""
        DROP TABLE IF EXISTS quality_records;
        DROP TABLE IF EXISTS deliveries;
        DROP TABLE IF EXISTS purchase_orders;
        DROP TABLE IF EXISTS suppliers;

        CREATE TABLE suppliers (
            supplier_id    TEXT PRIMARY KEY,
            supplier_name  TEXT,
            category       TEXT,
            country        TEXT,
            contact_name   TEXT,
            contact_email  TEXT,
            avl_status     TEXT
        );

        CREATE TABLE purchase_orders (
            po_number              TEXT PRIMARY KEY,
            supplier_id            TEXT,
            part_number            TEXT,
            part_category          TEXT,
            unit_cost              REAL,
            quantity               INTEGER,
            total_value            REAL,
            po_date                TEXT,
            committed_lead_time_days INTEGER,
            escalation_flag        INTEGER DEFAULT 0,
            FOREIGN KEY (supplier_id) REFERENCES suppliers(supplier_id)
        );

        CREATE TABLE deliveries (
            delivery_id              TEXT PRIMARY KEY,
            po_number                TEXT,
            committed_delivery_date  TEXT,
            actual_delivery_date     TEXT,
            received_qty             INTEGER,
            on_time                  INTEGER,
            in_full                  INTEGER,
            FOREIGN KEY (po_number) REFERENCES purchase_orders(po_number)
        );

        CREATE TABLE quality_records (
            inspection_id     TEXT PRIMARY KEY,
            delivery_id       TEXT,
            inspection_result TEXT,
            defect_qty        INTEGER,
            defect_type       TEXT,
            FOREIGN KEY (delivery_id) REFERENCES deliveries(delivery_id)
        );
    """)

    suppliers = [
        ("S001", "Delta Fabrication",  "Mechanical",  "USA",    "John Park",     "j.park@deltafab.com",    "Approved"),
        ("S002", "NovaDrive",          "Electrical",  "USA",    "Sara Chen",     "s.chen@novadrive.com",   "Approved"),
        ("S003", "CircuitPro",         "Electronics", "Mexico", "Luis Reyes",    "l.reyes@circuitpro.com", "Approved"),
        ("S004", "Apex Components",    "Mechanical",  "Canada", "Mike Johnson",  "m.j@apex.com",           "Conditional"),
        ("S005", "SealTech",           "Hardware",    "USA",    "Amy Wu",        "a.wu@sealtech.com",      "Approved"),
        ("S006", "CableTech",          "Electrical",  "China",  "Wei Zhang",     "w.zhang@cabletech.com",  "Conditional"),
        ("S007", "GlobalElec",         "Electronics", "Taiwan", "Kevin Lin",     "k.lin@globalelec.com",   "Approved"),
        ("S008", "FluidMaster",        "Hardware",    "USA",    "Tom Baker",     "t.baker@fluidmaster.com","Approved"),
    ]
    c.executemany("INSERT INTO suppliers VALUES (?,?,?,?,?,?,?)", suppliers)

    # OTIF profiles: (on_time_rate, in_full_rate, lead_time_variance_avg, quality_pass_rate)
    profiles = {
        "S001": (0.95, 0.97, -0.5, 0.97),
        "S002": (0.90, 0.93, 1.2,  0.94),
        "S003": (0.88, 0.90, 2.1,  0.92),
        "S004": (0.78, 0.82, 4.5,  0.88),
        "S005": (0.96, 0.98, -1.0, 0.99),
        "S006": (0.70, 0.75, 7.2,  0.85),
        "S007": (0.92, 0.94, 0.8,  0.95),
        "S008": (0.93, 0.96, 0.2,  0.98),
    }

    categories = {
        "S001": ("Mechanical",  ["PN-M001","PN-M002","PN-M003"]),
        "S002": ("Electrical",  ["PN-E001","PN-E002"]),
        "S003": ("Electronics", ["PN-X001","PN-X002"]),
        "S004": ("Mechanical",  ["PN-M004","PN-M005"]),
        "S005": ("Hardware",    ["PN-H001"]),
        "S006": ("Electrical",  ["PN-E003"]),
        "S007": ("Electronics", ["PN-X003","PN-X004"]),
        "S008": ("Hardware",    ["PN-H002","PN-H003"]),
    }

    po_num = 3000
    del_num = 5000
    ins_num = 7000

    for sid, (on_time_r, in_full_r, lt_var, quality_r) in profiles.items():
        cat, parts = categories[sid]
        n_pos = random.randint(12, 20)
        for _ in range(n_pos):
            po = f"PO-{po_num}"; po_num += 1
            part = random.choice(parts)
            uc   = round(random.uniform(5, 250), 2)
            qty  = random.randint(10, 200)
            tv   = round(uc * qty, 2)
            days_ago = random.randint(10, 365)
            po_date  = (TODAY - timedelta(days=days_ago)).strftime("%Y-%m-%d")
            clt  = random.randint(14, 45)
            esc  = 1 if random.random() > (0.85 if on_time_r > 0.85 else 0.65) else 0

            c.execute("INSERT INTO purchase_orders VALUES (?,?,?,?,?,?,?,?,?,?)",
                      (po, sid, part, cat, uc, qty, tv, po_date, clt, esc))

            # Delivery
            did = f"D-{del_num}"; del_num += 1
            commit_date = (datetime.strptime(po_date, "%Y-%m-%d") + timedelta(days=clt)).strftime("%Y-%m-%d")
            actual_var  = int(random.gauss(lt_var, 3))
            actual_date = (datetime.strptime(commit_date, "%Y-%m-%d") + timedelta(days=actual_var)).strftime("%Y-%m-%d")
            ot   = 1 if actual_var <= 0 else (1 if random.random() < on_time_r else 0)
            inf  = 1 if random.random() < in_full_r else 0
            recv = qty if inf else int(qty * random.uniform(0.7, 0.95))

            c.execute("INSERT INTO deliveries VALUES (?,?,?,?,?,?,?)",
                      (did, po, commit_date, actual_date, recv, ot, inf))

            # Quality
            iid    = f"I-{ins_num}"; ins_num += 1
            passed = random.random() < quality_r
            defects= 0 if passed else random.randint(1, int(recv * 0.08) + 1)
            dtype  = random.choice(["Dimensional","Surface","Functional","Documentation","Packaging"]) if not passed else None
            c.execute("INSERT INTO quality_records VALUES (?,?,?,?,?)",
                      (iid, did, "PASS" if passed else "FAIL", defects, dtype))

    conn.commit()
    conn.close()
    print("✅ Database seeded")

# ── RUN SQL QUERIES ───────────────────────────────────────────────────────────
def run_queries():
    conn = sqlite3.connect(DB_PATH)

    scorecard = pd.read_sql("""
        SELECT s.supplier_name, s.category, s.country, s.avl_status,
            COUNT(po.po_number) AS total_pos,
            ROUND(SUM(po.total_value),2) AS total_spend,
            ROUND(100.0*SUM(CASE WHEN d.on_time=1 AND d.in_full=1 THEN 1 ELSE 0 END)/NULLIF(COUNT(d.delivery_id),0),1) AS otif_pct,
            ROUND(100.0*SUM(CASE WHEN d.on_time=1 THEN 1 ELSE 0 END)/NULLIF(COUNT(d.delivery_id),0),1) AS on_time_pct,
            ROUND(AVG(JULIANDAY(d.actual_delivery_date)-JULIANDAY(po.po_date)),1) AS avg_lead_time_days,
            ROUND(AVG(JULIANDAY(d.actual_delivery_date)-JULIANDAY(d.committed_delivery_date)),1) AS avg_lt_variance_days,
            ROUND(100.0*SUM(CASE WHEN q.inspection_result='PASS' THEN 1 ELSE 0 END)/NULLIF(COUNT(q.inspection_id),0),1) AS quality_pass_pct,
            SUM(COALESCE(q.defect_qty,0)) AS total_defects,
            SUM(COALESCE(po.escalation_flag,0)) AS escalations
        FROM suppliers s
        LEFT JOIN purchase_orders po ON s.supplier_id=po.supplier_id
        LEFT JOIN deliveries d ON po.po_number=d.po_number
        LEFT JOIN quality_records q ON d.delivery_id=q.delivery_id
        GROUP BY s.supplier_name,s.category,s.country,s.avl_status
        ORDER BY otif_pct DESC
    """, conn)

    otif_trend = pd.read_sql("""
        SELECT s.supplier_name,
            STRFTIME('%Y-%m', d.actual_delivery_date) AS month,
            ROUND(100.0*SUM(CASE WHEN d.on_time=1 AND d.in_full=1 THEN 1 ELSE 0 END)/NULLIF(COUNT(*),0),1) AS otif_pct
        FROM suppliers s
        JOIN purchase_orders po ON s.supplier_id=po.supplier_id
        JOIN deliveries d ON po.po_number=d.po_number
        WHERE d.actual_delivery_date >= DATE('2026-04-04','-12 months')
        GROUP BY s.supplier_name, month
        ORDER BY s.supplier_name, month
    """, conn)

    quality = pd.read_sql("""
        SELECT s.supplier_name, s.category,
            COUNT(q.inspection_id) AS total_inspections,
            SUM(CASE WHEN q.inspection_result='PASS' THEN 1 ELSE 0 END) AS passed,
            SUM(CASE WHEN q.inspection_result='FAIL' THEN 1 ELSE 0 END) AS failed,
            ROUND(100.0*SUM(CASE WHEN q.inspection_result='PASS' THEN 1 ELSE 0 END)/NULLIF(COUNT(*),0),1) AS pass_rate_pct,
            SUM(COALESCE(q.defect_qty,0)) AS total_defects
        FROM suppliers s
        JOIN purchase_orders po ON s.supplier_id=po.supplier_id
        JOIN deliveries d ON po.po_number=d.po_number
        JOIN quality_records q ON d.delivery_id=q.delivery_id
        GROUP BY s.supplier_name,s.category
        ORDER BY pass_rate_pct ASC
    """, conn)

    spend = pd.read_sql("""
        SELECT s.supplier_name, s.category,
            COUNT(po.po_number) AS po_count,
            ROUND(SUM(po.total_value),2) AS total_spend
        FROM suppliers s
        JOIN purchase_orders po ON s.supplier_id=po.supplier_id
        GROUP BY s.supplier_name,s.category
        ORDER BY total_spend DESC
    """, conn)

    risk = pd.read_sql("""
        SELECT s.supplier_name, s.category,
            SUM(COALESCE(po.escalation_flag,0)) AS escalations,
            ROUND(100.0*SUM(CASE WHEN d.on_time=1 AND d.in_full=1 THEN 1 ELSE 0 END)/NULLIF(COUNT(d.delivery_id),0),1) AS otif_pct,
            ROUND(100.0*SUM(CASE WHEN q.inspection_result='PASS' THEN 1 ELSE 0 END)/NULLIF(COUNT(q.inspection_id),0),1) AS quality_pct
        FROM suppliers s
        JOIN purchase_orders po ON s.supplier_id=po.supplier_id
        JOIN deliveries d ON po.po_number=d.po_number
        LEFT JOIN quality_records q ON d.delivery_id=q.delivery_id
        GROUP BY s.supplier_name,s.category
        ORDER BY otif_pct ASC
    """, conn)

    conn.close()
    print("✅ SQL queries complete")
    return scorecard, otif_trend, quality, spend, risk

# ── BUILD EXCEL DASHBOARD ─────────────────────────────────────────────────────
def style_header_row(ws, row, cols, bg=DARK_NAVY):
    for col in range(1, cols+1):
        cell = ws.cell(row=row, column=col)
        cell.font      = hdr_font()
        cell.fill      = fill(bg)
        cell.alignment = center()
        cell.border    = border()

def style_data_row(ws, row, cols, bg=WHITE, bold=False):
    for col in range(1, cols+1):
        cell = ws.cell(row=row, column=col)
        cell.font      = body_font(bold=bold)
        cell.fill      = fill(bg)
        cell.alignment = left()
        cell.border    = border()

def write_df(ws, df, start_row, headers=None):
    if headers is None:
        headers = list(df.columns)
    for ci, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=ci).value = h
    style_header_row(ws, start_row, len(headers))

    for ri, (_, row_data) in enumerate(df.iterrows(), start_row+1):
        bg = GRAY_HDR if ri % 2 == 0 else WHITE
        for ci, val in enumerate(row_data, 1):
            ws.cell(row=ri, column=ci).value = val
        style_data_row(ws, ri, len(row_data), bg=bg)
    return start_row + len(df) + 1

def build_excel(scorecard, otif_trend, quality, spend, risk):
    wb = Workbook()
    wb.remove(wb.active)

    # ── Sheet 1: Summary Dashboard ────────────────────────────────────────
    ws = wb.create_sheet("📊 Summary Dashboard")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.row_dimensions[1].height = 8

    # Title banner
    ws.merge_cells("B2:L3")
    title = ws["B2"]
    title.value     = "SUPPLIER PERFORMANCE DASHBOARD"
    title.font      = Font(name="Arial", size=16, bold=True, color=WHITE)
    title.fill      = fill(DARK_NAVY)
    title.alignment = center()

    ws.merge_cells("B4:L4")
    sub = ws["B4"]
    sub.value     = f"Author: Nisarga Narasimhamurthy  |  Data as of: {TODAY.strftime('%B %d, %Y')}  |  Suppliers: {len(scorecard)}"
    sub.font      = Font(name="Arial", size=10, color=WHITE)
    sub.fill      = fill(MID_NAVY)
    sub.alignment = center()
    ws.row_dimensions[4].height = 18

    # KPI cards row
    kpi_labels = ["Avg OTIF %", "Avg Quality %", "Total Spend ($)", "Total POs", "Total Escalations", "Avg Lead Time (days)"]
    kpi_values = [
        f"{scorecard['otif_pct'].mean():.1f}%",
        f"{scorecard['quality_pass_pct'].mean():.1f}%",
        f"${scorecard['total_spend'].sum():,.0f}",
        f"{int(scorecard['total_pos'].sum())}",
        f"{int(scorecard['escalations'].sum())}",
        f"{scorecard['avg_lead_time_days'].mean():.1f}",
    ]
    kpi_cols = ["B","C","D","E","F","G"]
    kpi_colors = [LIGHT_BLUE, GREEN_BG, AMBER_BG, LIGHT_BLUE, RED_BG, GREEN_BG]
    kpi_txt_colors = ["0C447C", GREEN_TXT, AMBER_TXT, "0C447C", RED_TXT, GREEN_TXT]

    for i, (col, label, val, bg, tc) in enumerate(zip(kpi_cols, kpi_labels, kpi_values, kpi_colors, kpi_txt_colors)):
        # Label
        lc = ws[f"{col}6"]
        lc.value     = label
        lc.font      = Font(name="Arial", size=9, bold=True, color=tc)
        lc.fill      = fill(bg)
        lc.alignment = center()
        lc.border    = border()
        ws.row_dimensions[6].height = 16
        # Value
        vc = ws[f"{col}7"]
        vc.value     = val
        vc.font      = Font(name="Arial", size=14, bold=True, color=tc)
        vc.fill      = fill(bg)
        vc.alignment = center()
        vc.border    = border()
        ws.row_dimensions[7].height = 28
        ws.column_dimensions[col].width = 18

    # Scorecard mini table
    ws["B9"].value = "SUPPLIER SCORECARD SUMMARY"
    ws["B9"].font  = hdr_font(11, True, DARK_NAVY)

    headers = ["Supplier", "Category", "OTIF %", "On-Time %", "Quality %", "Avg LT (days)", "LT Variance", "Escalations", "AVL Status"]
    sc_cols = ["supplier_name","category","otif_pct","on_time_pct","quality_pass_pct","avg_lead_time_days","avg_lt_variance_days","escalations","avl_status"]
    write_df(ws, scorecard[sc_cols], 10, headers)

    # Conditional color on OTIF
    from openpyxl.formatting.rule import ColorScaleRule
    otif_col = "D"
    ws.conditional_formatting.add(
        f"{otif_col}11:{otif_col}{10+len(scorecard)}",
        ColorScaleRule(start_type="num", start_value=60, start_color="FCEBEB",
                       mid_type="num",   mid_value=85,   mid_color="FAEEDA",
                       end_type="num",   end_value=100,  end_color="E1F5EE")
    )

    # ── Sheet 2: OTIF Trend ───────────────────────────────────────────────
    ws2 = wb.create_sheet("📈 OTIF Trend")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 3
    ws2["B2"].value = "OTIF TREND — LAST 12 MONTHS"
    ws2["B2"].font  = hdr_font(12, True, DARK_NAVY)

    pivot = otif_trend.pivot(index="supplier_name", columns="month", values="otif_pct").reset_index()
    pivot.columns.name = None
    write_df(ws2, pivot, 4)
    for col in range(1, len(pivot.columns)+1):
        ws2.column_dimensions[get_column_letter(col+1)].width = 14

    # Line chart
    chart = LineChart()
    chart.title  = "OTIF % by Supplier — Monthly Trend"
    chart.y_axis.title = "OTIF %"
    chart.x_axis.title = "Month"
    chart.style  = 10
    chart.height = 14
    chart.width  = 28

    months = list(pivot.columns[1:])
    for i in range(len(pivot)):
        row_start = 5 + i
        data_ref = Reference(ws2, min_col=3, max_col=2+len(months), min_row=row_start, max_row=row_start)
        from openpyxl.chart import Series
        series = Series(data_ref, title=pivot.iloc[i]["supplier_name"])
        chart.series.append(series)

    ws2.add_chart(chart, "B" + str(6 + len(pivot) + 2))

    # ── Sheet 3: Quality ──────────────────────────────────────────────────
    ws3 = wb.create_sheet("🔍 Quality Report")
    ws3.sheet_view.showGridLines = False
    ws3.column_dimensions["A"].width = 3
    ws3["B2"].value = "QUALITY & INSPECTION REPORT"
    ws3["B2"].font  = hdr_font(12, True, DARK_NAVY)
    write_df(ws3, quality, 4, ["Supplier","Category","Inspections","Passed","Failed","Pass Rate %","Total Defects"])
    for col in "BCDEFGH":
        ws3.column_dimensions[col].width = 18
    ws3.conditional_formatting.add(
        f"G5:G{4+len(quality)}",
        ColorScaleRule(start_type="min", start_color="E1F5EE",
                       end_type="max",   end_color="FCEBEB")
    )

    # ── Sheet 4: Spend ────────────────────────────────────────────────────
    ws4 = wb.create_sheet("💰 Spend Analysis")
    ws4.sheet_view.showGridLines = False
    ws4.column_dimensions["A"].width = 3
    ws4["B2"].value = "SPEND ANALYSIS BY SUPPLIER"
    ws4["B2"].font  = hdr_font(12, True, DARK_NAVY)
    spend["spend_pct"] = (spend["total_spend"] / spend["total_spend"].sum() * 100).round(1)
    write_df(ws4, spend, 4, ["Supplier","Category","PO Count","Total Spend ($)","Spend %"])
    for col in "BCDEF":
        ws4.column_dimensions[col].width = 20

    # Bar chart
    chart2       = BarChart()
    chart2.type  = "col"
    chart2.title = "Total Spend by Supplier"
    chart2.y_axis.title = "Spend ($)"
    chart2.style = 10
    chart2.height = 14
    chart2.width  = 22
    data2  = Reference(ws4, min_col=5, max_col=5, min_row=4, max_row=4+len(spend))
    cats2  = Reference(ws4, min_col=2, max_row=4+len(spend), min_row=5)
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats2)
    ws4.add_chart(chart2, "B" + str(6 + len(spend) + 2))

    # ── Sheet 5: Risk Flags ───────────────────────────────────────────────
    ws5 = wb.create_sheet("🚨 Risk Flags")
    ws5.sheet_view.showGridLines = False
    ws5.column_dimensions["A"].width = 3
    ws5["B2"].value = "SUPPLIER RISK FLAGS"
    ws5["B2"].font  = hdr_font(12, True, DARK_NAVY)

    risk["risk_flag"] = risk.apply(lambda r:
        "🔴 HIGH RISK"   if r["escalations"] >= 3 or r["otif_pct"] < 75 else
        ("🟡 MONITOR"    if r["escalations"] >= 1 or r["otif_pct"] < 88 else
         "🟢 OK"), axis=1)
    write_df(ws5, risk, 4, ["Supplier","Category","Escalations","OTIF %","Quality %","Risk Flag"])
    for col in "BCDEFG":
        ws5.column_dimensions[col].width = 20

    # Color risk flag column
    for ri, (_, row_data) in enumerate(risk.iterrows(), 5):
        flag = row_data["risk_flag"]
        cell = ws5.cell(row=ri, column=7)
        if "HIGH" in flag:
            cell.fill = fill(RED_BG)
            cell.font = body_font(color=RED_TXT, bold=True)
        elif "MONITOR" in flag:
            cell.fill = fill(AMBER_BG)
            cell.font = body_font(color=AMBER_TXT, bold=True)
        else:
            cell.fill = fill(GREEN_BG)
            cell.font = body_font(color=GREEN_TXT, bold=True)

    # ── Sheet 6: Raw Data ─────────────────────────────────────────────────
    ws6 = wb.create_sheet("📋 Raw Data")
    ws6.sheet_view.showGridLines = False
    ws6.column_dimensions["A"].width = 3
    ws6["B2"].value = "RAW SCORECARD DATA (SQL OUTPUT)"
    ws6["B2"].font  = hdr_font(12, True, DARK_NAVY)
    write_df(ws6, scorecard, 4)
    for col_idx in range(1, len(scorecard.columns)+2):
        ws6.column_dimensions[get_column_letter(col_idx)].width = 20

    wb.save(OUT_PATH)
    print(f"✅ Excel dashboard saved → {OUT_PATH}")

# ── MAIN ──────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("\n" + "="*60)
    print("  SUPPLIER PERFORMANCE DASHBOARD BUILDER")
    print("  Nisarga Narasimhamurthy | Supply Chain Portfolio")
    print("="*60 + "\n")

    seed_database()
    scorecard, otif_trend, quality, spend, risk = run_queries()

    print(f"\n📊 Scorecard: {len(scorecard)} suppliers")
    print(f"   Avg OTIF     : {scorecard['otif_pct'].mean():.1f}%")
    print(f"   Avg Quality  : {scorecard['quality_pass_pct'].mean():.1f}%")
    print(f"   Total Spend  : ${scorecard['total_spend'].sum():,.0f}")
    print(f"   Escalations  : {int(scorecard['escalations'].sum())}")

    build_excel(scorecard, otif_trend, quality, spend, risk)

    print("\n📁 FILES GENERATED:")
    print("   supplier_data.db                    — SQLite database")
    print("   supplier_performance_dashboard.xlsx — Excel dashboard (6 sheets)")
    print("\n" + "="*60 + "\n")
